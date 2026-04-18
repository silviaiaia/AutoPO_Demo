# Append parsed rows to the Open Order workbook.
# Production uses xlwings against a live Excel session (Windows-only); the
# demo uses openpyxl against a file on disk. Same row/column logic either way.

from __future__ import annotations

from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from autopo.config import DATE_COLUMNS, STANDARD_COLUMNS


_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def _default_headers() -> List[str]:
    order = [
        "upload", "sales_doc_item", "soldto_abbr", "shipto_abbr",
        "customer_ref_date", "soldto", "shipto", "customer_ref",
        "customer_po_item", "material", "module_material",
        "customer_material", "order_qty", "unit_price", "currency",
        "crd", "etd", "remark",
    ]
    return [STANDARD_COLUMNS[k] for k in order]


def ensure_workbook(path: str | Path, sheet_name: str = "OpenOrder") -> Workbook:
    """Load an existing workbook or create a fresh one with headers."""
    path = Path(path)
    if path.exists():
        wb = load_workbook(path)
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        return wb

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    _write_header(ws)
    return wb


def _write_header(ws) -> None:
    headers = _default_headers()
    for idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    for idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 18
    ws.freeze_panes = "A2"


def append_rows(
    workbook_path: str | Path,
    rows: Iterable[dict],
    sheet_name: str = "OpenOrder",
) -> int:
    """Append rows to the target sheet; return count written."""
    path = Path(workbook_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = ensure_workbook(path, sheet_name)
    ws = wb[sheet_name]

    # Use the existing header row if there is one.
    headers = [c.value for c in ws[1] if c.value] or _default_headers()
    # Fill any missing cells so the header row is always complete.
    for idx, name in enumerate(_default_headers(), start=1):
        if ws.cell(row=1, column=idx).value is None:
            ws.cell(row=1, column=idx, value=name)
            ws.cell(row=1, column=idx).font = _HEADER_FONT
            ws.cell(row=1, column=idx).fill = _HEADER_FILL

    start_row = ws.max_row + 1
    written = 0
    for row in rows:
        for idx, header in enumerate(headers, start=1):
            value = row.get(header, "")
            cell = ws.cell(row=start_row + written, column=idx, value=value)
            if header in DATE_COLUMNS and value:
                cell.number_format = "YYYY/MM/DD"
        written += 1

    wb.save(path)
    return written
